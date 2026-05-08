# -*- coding: utf-8 -*-
"""
分析相关API蓝图
处理数据分析、报告生成等功能
"""

from flask import Blueprint, request, jsonify, send_file
import os

analytics_bp = Blueprint('analytics', __name__, url_prefix='/api/analytics')


@analytics_bp.route('/export')
def export_analytics():
    """导出分析报告"""
    try:
        from src.web.service_manager import service_manager
        from src.core.query_optimizer import query_optimizer
        from openpyxl import Workbook
        from openpyxl.styles import Font

        # 生成Excel报告（使用数据库真实数据）
        analytics = query_optimizer.get_analytics_optimized(30)

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "分析报告"

        # 添加标题
        ws['A1'] = 'AI Helpdesk 分析报告'
        ws['A1'].font = Font(size=16, bold=True)

        # 添加工单统计
        ws['A3'] = '工单统计'
        ws['A3'].font = Font(bold=True)
        ws['A4'] = '总工单数'
        ws['B4'] = analytics['workorders']['total']
        ws['A5'] = '待处理'
        ws['B5'] = analytics['workorders']['open']
        ws['A6'] = '已解决'
        ws['B6'] = analytics['workorders']['resolved']

        # 保存文件
        report_path = 'uploads/analytics_report.xlsx'
        os.makedirs('uploads', exist_ok=True)
        wb.save(report_path)

        return send_file(report_path, as_attachment=True, download_name='analytics_report.xlsx')

    except Exception as e:
        return jsonify({"error": str(e)}), 500
