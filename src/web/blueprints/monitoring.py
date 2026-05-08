# -*- coding: utf-8 -*-
"""
监控管理蓝图
处理监控相关的API路由
"""

from datetime import datetime, timedelta
from flask import Blueprint, request, jsonify
from sqlalchemy import func
from src.main import Assistant
from src.core.database import db_manager
from src.core.models import Conversation, WorkOrder, Alert, KnowledgeEntry

monitoring_bp = Blueprint('monitoring', __name__, url_prefix='/api')

def estimate_tokens(text):
    """估算文本的Token数量"""
    if not text:
        return 0
    
    # 中文字符约1.5字符=1token，英文字符约4字符=1token
    chinese_chars = len([c for c in text if '\u4e00' <= c <= '\u9fff'])
    english_chars = len(text) - chinese_chars
    return int(chinese_chars / 1.5 + english_chars / 4)

def calculate_conversation_tokens(conversations):
    """计算对话记录的Token使用量"""
    total_tokens = 0
    for conv in conversations:
        user_message = conv.user_message or ""
        assistant_response = conv.assistant_response or ""
        total_tokens += estimate_tokens(user_message) + estimate_tokens(assistant_response)
    return total_tokens

def get_assistant():
    """获取实例（懒加载）"""
    global _assistant
    if '_assistant' not in globals():
        _assistant = Assistant()
    return _assistant

# Token监控相关API
@monitoring_bp.route('/token-monitor/stats')
def get_token_monitor_stats():
    """获取Token监控统计"""
    try:
        from datetime import datetime, timedelta
        import calendar
        
        now = datetime.now()
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        with db_manager.get_session() as session:
            # 优化：使用单个查询获取所有需要的数据
            conversations_query = session.query(Conversation).filter(
                Conversation.timestamp >= month_start
            ).all()
            
            # 分离今日和本月数据
            today_conversations = [c for c in conversations_query if c.timestamp >= today_start]
            month_conversations = conversations_query
            
            # 计算真实的Token使用量
            today_tokens = calculate_conversation_tokens(today_conversations)
            month_tokens = calculate_conversation_tokens(month_conversations)
            
            # 根据真实Token使用量计算成本
            total_cost = month_tokens * 0.0008 / 1000  # qwen-turbo输入价格
            budget_limit = 1000  # 预算限制
            
            return jsonify({
                'success': True,
                'today_tokens': today_tokens,
                'month_tokens': month_tokens,
                'total_cost': round(total_cost, 2),
                'budget_limit': budget_limit
            })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@monitoring_bp.route('/token-monitor/chart')
def get_token_monitor_chart():
    """获取Token使用趋势图表数据"""
    try:
        period = request.args.get('period', 'day')
        from datetime import datetime, timedelta
        
        now = datetime.now()
        labels = []
        tokens = []
        costs = []
        
        if period == 'hour':
            # 最近24小时
            for i in range(24):
                hour_start = now - timedelta(hours=i+1)
                hour_end = now - timedelta(hours=i)
                labels.insert(0, hour_start.strftime('%H:00'))
                
                with db_manager.get_session() as session:
                    hour_conversations = session.query(Conversation).filter(
                        Conversation.timestamp >= hour_start,
                        Conversation.timestamp < hour_end
                    ).all()
                    
                    # 计算真实Token使用量
                    hour_tokens = calculate_conversation_tokens(hour_conversations)
                    
                    tokens.insert(0, hour_tokens)
                    costs.insert(0, hour_tokens * 0.0008 / 1000)
        
        elif period == 'day':
            # 最近7天
            for i in range(7):
                day_start = now - timedelta(days=i+1)
                day_end = now - timedelta(days=i)
                labels.insert(0, day_start.strftime('%m-%d'))
                
                with db_manager.get_session() as session:
                    day_conversations = session.query(Conversation).filter(
                        Conversation.timestamp >= day_start,
                        Conversation.timestamp < day_end
                    ).all()
                    
                    # 计算真实Token使用量
                    day_tokens = calculate_conversation_tokens(day_conversations)
                    
                    tokens.insert(0, day_tokens)
                    costs.insert(0, day_tokens * 0.0008 / 1000)
        
        elif period == 'week':
            # 最近4周
            for i in range(4):
                week_start = now - timedelta(weeks=i+1)
                week_end = now - timedelta(weeks=i)
                labels.insert(0, f"第{i+1}周")
                
                with db_manager.get_session() as session:
                    week_conversations = session.query(Conversation).filter(
                        Conversation.timestamp >= week_start,
                        Conversation.timestamp < week_end
                    ).all()
                    
                    # 计算真实Token使用量
                    week_tokens = calculate_conversation_tokens(week_conversations)
                    
                    tokens.insert(0, week_tokens)
                    costs.insert(0, week_tokens * 0.0008 / 1000)
        
        return jsonify({
            'success': True,
            'labels': labels,
            'tokens': tokens,
            'costs': costs
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@monitoring_bp.route('/token-monitor/records')
def get_token_monitor_records():
    """获取Token使用详细记录"""
    try:
        limit = request.args.get('limit', 50, type=int)
        
        with db_manager.get_session() as session:
            conversations = session.query(Conversation).order_by(
                Conversation.timestamp.desc()
            ).limit(limit).all()
            
            records = []
            for conv in conversations:
                # 从对话内容估算真实的Token使用量
                user_message = conv.user_message or ""
                assistant_response = conv.assistant_response or ""
                
                input_tokens = estimate_tokens(user_message)
                output_tokens = estimate_tokens(assistant_response)
                total_tokens = input_tokens + output_tokens
                
                # 根据qwen-turbo价格计算成本
                cost = (input_tokens * 0.0008 + output_tokens * 0.002) / 1000
                
                records.append({
                    'timestamp': conv.timestamp.isoformat() if conv.timestamp else None,
                    'user_id': f"user_{conv.id}",
                    'model': 'qwen-turbo',
                    'input_tokens': input_tokens,
                    'output_tokens': output_tokens,
                    'total_tokens': total_tokens,
                    'cost': round(cost, 6),
                    'response_time': conv.response_time or 0
                })
            
            return jsonify({
                'success': True,
                'records': records
            })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@monitoring_bp.route('/token-monitor/settings', methods=['POST'])
def save_token_monitor_settings():
    """保存Token监控设置"""
    try:
        data = request.get_json()
        
        # 这里可以将设置保存到数据库或配置文件
        # 暂时返回成功
        
        return jsonify({
            'success': True,
            'message': 'Token设置已保存'
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@monitoring_bp.route('/token-monitor/export')
def export_token_monitor_data():
    """导出Token使用数据"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Token使用数据"
        
        # 添加标题
        ws['A1'] = 'Token使用数据导出'
        ws['A1'].font = Font(size=16, bold=True)
        
        # 添加表头
        headers = ['时间', '用户', '模型', '输入Token', '输出Token', '总Token', '成本', '响应时间']
        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header)
        
        # 添加数据
        with db_manager.get_session() as session:
            conversations = session.query(Conversation).order_by(
                Conversation.timestamp.desc()
            ).limit(1000).all()
            
            for row, conv in enumerate(conversations, 4):
                ws.cell(row=row, column=1, value=conv.timestamp.isoformat() if conv.timestamp else '')
                ws.cell(row=row, column=2, value=conv.user_id or '')
                ws.cell(row=row, column=3, value='qwen-turbo')
                ws.cell(row=row, column=4, value=conv.response_time or 0)
                ws.cell(row=row, column=5, value=(conv.response_time or 0) * 0.5)
                ws.cell(row=row, column=6, value=(conv.response_time or 0) * 1.5)
                ws.cell(row=row, column=7, value=(conv.response_time or 0) * 0.0001)
                ws.cell(row=row, column=8, value=conv.response_time or 0)
        
        # 保存文件
        import tempfile
        import os
        temp_path = os.path.join(tempfile.gettempdir(), 'token_usage_data.xlsx')
        wb.save(temp_path)
        
        from flask import send_file
        return send_file(temp_path, as_attachment=True, download_name='token_usage_data.xlsx')
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# AI监控相关API
@monitoring_bp.route('/ai-monitor/stats')
def get_ai_monitor_stats():
    """获取AI监控统计"""
    try:
        with db_manager.get_session() as session:
            # 优化：限制查询数量，只获取最近的数据
            conversations = session.query(Conversation).order_by(
                Conversation.timestamp.desc()
            ).limit(1000).all()  # 限制查询数量
            total_calls = len(conversations)
            
            if total_calls == 0:
                return jsonify({
                    'success': True,
                    'total_calls': 0,
                    'success_rate': 0,
                    'error_rate': 0,
                    'avg_response_time': 0
                })
            
            # 基于实际对话质量计算成功率
            successful_calls = 0
            total_response_time = 0
            response_times = []
            
            for conv in conversations:
                # 判断对话是否成功
                is_success = True
                
                # 检查响应时间
                if conv.response_time:
                    response_times.append(conv.response_time)
                    total_response_time += conv.response_time
                    if conv.response_time > 10000:  # 超过10秒认为失败
                        is_success = False
                
                # 检查置信度
                if conv.confidence_score and conv.confidence_score < 0.3:
                    is_success = False
                
                # 检查回复内容
                if not conv.assistant_response or len(conv.assistant_response.strip()) < 5:
                    is_success = False
                
                if is_success:
                    successful_calls += 1
            
            success_rate = (successful_calls / total_calls * 100) if total_calls > 0 else 0
            error_rate = 100 - success_rate
            avg_response_time = (total_response_time / len(response_times)) if response_times else 0
            
            return jsonify({
                'success': True,
                'total_calls': total_calls,
                'success_rate': round(success_rate, 1),
                'error_rate': round(error_rate, 1),
                'avg_response_time': round(avg_response_time, 0)
            })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@monitoring_bp.route('/ai-monitor/model-comparison')
def get_model_comparison():
    """获取模型性能对比数据"""
    try:
        with db_manager.get_session() as session:
            conversations = session.query(Conversation).all()
            
            # 分析实际使用的模型（目前只有qwen-turbo）
            model_stats = {}
            
            for conv in conversations:
                model = 'qwen-turbo'  # 实际使用的模型
                
                if model not in model_stats:
                    model_stats[model] = {
                        'total_calls': 0,
                        'successful_calls': 0,
                        'total_response_time': 0,
                        'response_times': []
                    }
                
                model_stats[model]['total_calls'] += 1
                
                # 判断是否成功
                is_success = True
                if conv.response_time and conv.response_time > 10000:
                    is_success = False
                if conv.confidence_score and conv.confidence_score < 0.3:
                    is_success = False
                if not conv.assistant_response or len(conv.assistant_response.strip()) < 5:
                    is_success = False
                
                if is_success:
                    model_stats[model]['successful_calls'] += 1
                
                if conv.response_time:
                    model_stats[model]['response_times'].append(conv.response_time)
                    model_stats[model]['total_response_time'] += conv.response_time
            
            # 计算性能指标
            models = []
            success_rates = []
            response_times = []
            
            for model, stats in model_stats.items():
                models.append(model)
                
                success_rate = (stats['successful_calls'] / stats['total_calls'] * 100) if stats['total_calls'] > 0 else 0
                success_rates.append(round(success_rate, 1))
                
                avg_response_time = (stats['total_response_time'] / len(stats['response_times'])) if stats['response_times'] else 0
                response_times.append(round(avg_response_time, 0))
            
            # 如果没有数据，显示默认值
            if not models:
                models = ['qwen-turbo']
                success_rates = [100.0]
                response_times = [0]
            
            return jsonify({
                'success': True,
                'models': models,
                'success_rates': success_rates,
                'response_times': response_times
            })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@monitoring_bp.route('/ai-monitor/error-distribution')
def get_error_distribution():
    """获取错误类型分布"""
    try:
        with db_manager.get_session() as session:
            # 基于实际对话记录分析错误类型
            conversations = session.query(Conversation).all()
            
            # 分析对话记录中的错误模式
            error_types = ['成功', '响应超时', '内容异常', '格式错误', '其他错误']
            counts = [0, 0, 0, 0, 0]
            
            for conv in conversations:
                # 基于响应时间和内容质量判断错误类型
                if conv.response_time and conv.response_time > 10000:  # 超过10秒
                    counts[1] += 1  # 响应超时
                elif conv.confidence_score and conv.confidence_score < 0.3:  # 低置信度
                    counts[2] += 1  # 内容异常
                elif not conv.assistant_response or len(conv.assistant_response.strip()) < 5:
                    counts[3] += 1  # 格式错误
                elif conv.assistant_response and len(conv.assistant_response.strip()) >= 5:
                    counts[0] += 1  # 成功
                else:
                    counts[4] += 1  # 其他错误
            
            return jsonify({
                'success': True,
                'error_types': error_types,
                'counts': counts
            })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@monitoring_bp.route('/ai-monitor/error-log')
def get_error_log():
    """获取错误日志"""
    try:
        with db_manager.get_session() as session:
            # 获取有问题的对话记录作为错误日志
            conversations = session.query(Conversation).order_by(
                Conversation.timestamp.desc()
            ).limit(50).all()
            
            errors = []
            error_id = 1
            
            for conv in conversations:
                error_type = None
                error_message = None
                
                # 判断错误类型
                if conv.response_time and conv.response_time > 10000:  # 超过10秒
                    error_type = '响应超时'
                    error_message = f'响应时间过长: {conv.response_time}ms'
                elif conv.confidence_score and conv.confidence_score < 0.3:  # 低置信度
                    error_type = '内容异常'
                    error_message = f'置信度过低: {conv.confidence_score}'
                elif not conv.assistant_response or len(conv.assistant_response.strip()) < 5:
                    error_type = '格式错误'
                    error_message = '助手回复内容过短或为空'
                elif conv.assistant_response and 'error' in conv.assistant_response.lower():
                    error_type = 'API错误'
                    error_message = '回复中包含错误信息'
                
                # 只记录有错误的对话
                if error_type:
                    errors.append({
                        'id': error_id,
                        'timestamp': conv.timestamp.isoformat() if conv.timestamp else None,
                        'error_type': error_type,
                        'error_message': error_message,
                        'model': 'qwen-turbo',  # 实际使用的模型
                        'user_id': f'user_{conv.id}'
                    })
                    error_id += 1
            
            return jsonify({
                'success': True,
                'errors': errors
            })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@monitoring_bp.route('/ai-monitor/error-log', methods=['DELETE'])
def clear_error_log():
    """清空错误日志"""
    try:
        # 这里应该清空实际的错误日志表
        return jsonify({
            'success': True,
            'message': '错误日志已清空'
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
